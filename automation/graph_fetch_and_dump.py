#!/usr/bin/env python3
"""
Fetches "Showrooms Revamp Timeline.xlsx" directly from Microsoft Graph using
app-only (client-credentials) auth, and re-serializes the four sheets
refresh_dashboard.py needs into the same tab-separated "## Sheet: <name>"
text format the Microsoft 365 connector normally produces — so the existing,
already-proven refresh_dashboard.py (copied unmodified into this folder) can
consume it exactly as it does in the interactive/OneDrive-connector workflow.

This script is the piece that only works once IT has completed the Azure AD
app registration described in docs/GITHUB_ACTIONS_SETUP.md. It reads three
secrets from the environment (set as GitHub Actions repository secrets):

    AZURE_TENANT_ID
    AZURE_CLIENT_ID
    AZURE_CLIENT_SECRET

The workbook's drive/item identifiers below are NOT secrets — they are just
Microsoft Graph object IDs (the same ones the Microsoft 365 connector already
uses) and are safe to keep in source control. They only resolve to anything
once presented with a valid, authorized access token.

Usage: graph_fetch_and_dump.py <output_path>
"""
import os
import sys
import datetime as dt
from io import BytesIO

import requests
import openpyxl

# Drive item for "Showrooms Revamp Timeline.xlsx" (owner: mina_faham_com),
# the same file the hourly Microsoft 365 connector job reads.
DRIVE_ID = "b!fH2okKMc606AhlGPvxRbqQIGlG9GCfNCn8EedNT7pWp6O0pSKU0HRo7S4etT9Wdk"
ITEM_ID = "01QNMY335JO45ZYDVEVBCILPSTOHD2QBSI"

# Sheets refresh_dashboard.py looks for, matched by substring — mirrors its
# own next(k for k in sheets if 'V2' in k ...) logic. If a new weekly MOM
# sheet (e.g. "3 Sep MOM") replaces one of these, both this list AND
# refresh_dashboard.py's matching in main() need updating — same maintenance
# point that already exists in the interactive/OneDrive workflow.
NEEDED_SHEET_MATCH = ['V2', 'V1', '20 Aug', '27 Aug']


def get_token():
    tenant = os.environ['AZURE_TENANT_ID']
    client_id = os.environ['AZURE_CLIENT_ID']
    client_secret = os.environ['AZURE_CLIENT_SECRET']
    resp = requests.post(
        f'https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token',
        data={
            'client_id': client_id,
            'client_secret': client_secret,
            'grant_type': 'client_credentials',
            'scope': 'https://graph.microsoft.com/.default',
        },
        timeout=30,
    )
    if resp.status_code != 200:
        raise SystemExit(
            f'Token request failed ({resp.status_code}): {resp.text[:500]}\n'
            f'Check AZURE_TENANT_ID / AZURE_CLIENT_ID / AZURE_CLIENT_SECRET '
            f'secrets and that admin consent was granted for the app.'
        )
    return resp.json()['access_token']


def download_workbook(token):
    url = f'https://graph.microsoft.com/v1.0/drives/{DRIVE_ID}/items/{ITEM_ID}/content'
    resp = requests.get(url, headers={'Authorization': f'Bearer {token}'}, timeout=60)
    if resp.status_code != 200:
        raise SystemExit(
            f'Graph file download failed ({resp.status_code}): {resp.text[:500]}\n'
            f'Check the app has Files.Read.All (application, admin-consented) '
            f'and that DRIVE_ID/ITEM_ID in this script are still correct.'
        )
    return BytesIO(resp.content)


def fmt_cell(v):
    """Render one cell the way the Microsoft 365 connector's text dump does,
    closely enough for refresh_dashboard.py's parsers: dates as DD.MM.YY
    (parse_dotted() reads the real year straight off the cell, so this is
    actually more robust than the connector's mixed D-Mon/DD.MM.YY dump,
    which depends on refresh_dashboard.py's hardcoded YEAR constant for any
    date printed without an explicit year)."""
    if v is None:
        return ''
    if isinstance(v, (dt.datetime, dt.date)):
        return f'{v:%d.%m.%y}'
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    return str(v).strip()


def dump_sheet(ws):
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = [fmt_cell(v) for v in row]
        if not any(cells):
            continue  # fully-blank rows are dropped — refresh_dashboard.py's
            # rows_of() drops them too (as literal blank lines, or as the
            # connector's own "[N empty rows]" annotations), so the resulting
            # `rows` list it works from is identical either way.
        while cells and cells[-1] == '':
            cells.pop()
        lines.append('\t'.join(cells))
    return '\n'.join(lines)


def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else '/tmp/raw_sheet_dump.txt'
    token = get_token()
    wb = openpyxl.load_workbook(download_workbook(token), data_only=True)

    chosen = {}
    for match in NEEDED_SHEET_MATCH:
        hit = next((t for t in wb.sheetnames if match in t), None)
        if not hit:
            raise SystemExit(
                f'Could not find a sheet matching "{match}" in {wb.sheetnames}. '
                f'The workbook structure may have changed — this needs a human look.'
            )
        chosen[match] = hit

    out = ['Workbook: fetched live via Microsoft Graph (app-only). Cell values are tab-separated rows.']
    for match in NEEDED_SHEET_MATCH:
        name = chosen[match]
        out.append(f'\n## Sheet: {name}')
        out.append(dump_sheet(wb[name]))

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(out) + '\n')

    header_count = sum(1 for line in open(out_path, encoding='utf-8') if line.startswith('## Sheet: '))
    print(f'OK — wrote {out_path} with {header_count} sheet(s): {list(chosen.values())}')


if __name__ == '__main__':
    main()
